"""The four operator reports (client spec) + CSV/Excel export.

  1. Attendance      — per person: First-In, Last-Out, Duration (per day in range).
  2. Group           — per group: Headcount, Attendance Compliance %.
  3. Entry/Exit Mismatch — transit sessions: unpaired (open/overdue) vs resolved.
  4. Unknown Attempts — face_unknown events: count + snapshots.

Each report has a JSON endpoint (for the UI table) and the same data is exportable as
CSV or XLSX via ?format=csv|xlsx. The four are deliberately fixed — the old generic
"summary" stays for back-compat but the UI uses these.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional
from urllib.parse import parse_qs, urlparse

from fastapi import APIRouter, Depends, Query
from fastapi.responses import JSONResponse, Response
from sqlalchemy import and_, func, select

import config
from db import session
from db.models import FRSAttendance, FRSEvent, FRSGroup, FRSPerson, TransitSession
from deps import require_service_token, allowed_camera_ids
from schemas import iso, naive

router = APIRouter(tags=["reports"])

REPORTS = ("attendance", "group", "mismatch", "unknown")

_SAFE = re.compile(r"^[A-Za-z0-9\-_]+$")


def _snapshot_file(value: str) -> Optional[Path]:
    """Resolve a stored snapshot reference (a '/snapshot?key=live:<id>' path, a bare
    'live:<id>' key, or a relative photo storage_key) to an on-disk image path, or
    None. Used to embed the actual image into XLSX exports."""
    if not value:
        return None
    key = value
    if value.startswith("/snapshot") or value.startswith("http"):
        q = parse_qs(urlparse(value).query)
        key = (q.get("key") or [""])[0]
    for prefix in ("live:", "ingest:"):
        if key.startswith(prefix):
            name = key[len(prefix):]
            if not _SAFE.match(name):
                return None
            p = config.DATA_PATH / "snapshots" / f"{name}.jpg"
            return p if p.exists() else None
    # else treat as a relative path under DATA_PATH (best effort, no traversal)
    if ".." in key or key.startswith("/"):
        return None
    p = config.DATA_PATH / key
    return p if p.exists() else None


# ── export helpers ─────────────────────────────────────────────────────────
def _csv(columns: list[str], rows: list[dict]) -> Response:
    # CSV is plain text — it can't hold images, so drop the snapshot column entirely
    # (an image lives only in the XLSX export).
    cols = [c for c in columns if c != "snapshot"]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    w.writeheader()
    for r in rows:
        w.writerow({c: r.get(c, "") for c in cols})
    return Response(buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=report.csv"})


def _xlsx(columns: list[str], rows: list[dict], title: str = "Report") -> Response:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    has_snap = "snapshot" in columns
    ws.append([c.replace("_", " ").title() for c in columns])
    for c in ws[1]:
        c.font = Font(bold=True)

    THUMB = 56  # px — embedded face thumbnail size
    for ri, r in enumerate(rows, start=2):
        ws.append([("" if c == "snapshot" else r.get(c, "")) for c in columns])
        if not has_snap:
            continue
        src = _snapshot_file(r.get("snapshot") or "")
        if not src:
            continue
        try:
            img = XLImage(str(src))
            img.width = img.height = THUMB
            col_letter = get_column_letter(columns.index("snapshot") + 1)
            ws.row_dimensions[ri].height = THUMB * 0.78  # pt
            ws.add_image(img, f"{col_letter}{ri}")
        except Exception:  # noqa: BLE001
            pass

    if has_snap:
        ws.column_dimensions[get_column_letter(columns.index("snapshot") + 1)].width = 10
    for c in ws[1]:
        c.alignment = Alignment(vertical="center")

    bio = io.BytesIO()
    wb.save(bio)
    return Response(
        bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=report.xlsx"})


def _respond(columns, rows, fmt, title):
    fmt = (fmt or "json").lower()
    if fmt == "csv":
        return _csv(columns, rows)
    if fmt in ("xlsx", "excel"):
        return _xlsx(columns, rows, title)
    return JSONResponse({"columns": columns, "items": rows, "total": len(rows)})


def _fmt_duration(seconds: Optional[float]) -> str:
    if not seconds or seconds < 0:
        return "—"
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    return f"{h}h {m}m"


# ── 1. Attendance: First-In, Last-Out, Duration ────────────────────────────
@router.get("/reports/attendance")
def report_attendance(day_from: str = Query(...), day_to: str = Query(...),
                      format: str = Query("json"),
                      _: None = Depends(require_service_token),
                      allowed: Optional[list[str]] = Depends(allowed_camera_ids)) -> Response:
    columns = ["snapshot", "day", "person_name", "first_in", "last_out", "duration"]
    with session() as s:
        conds = [FRSAttendance.day_key >= day_from, FRSAttendance.day_key <= day_to]
        if allowed is not None:
            if not allowed:
                return _respond(columns, [], format, "Attendance")
            conds.append(FRSAttendance.camera_id.in_(allowed))
        stmt = (select(
            FRSAttendance.day_key, FRSAttendance.person_id, FRSPerson.full_name,
            FRSAttendance.check_in_at, FRSAttendance.check_out_at,
            FRSAttendance.check_in_snapshot, FRSAttendance.check_out_snapshot,
        ).outerjoin(FRSPerson, FRSPerson.id == FRSAttendance.person_id)
         .where(and_(*conds))
         .order_by(FRSAttendance.day_key.desc(), FRSPerson.full_name))
        rows = []
        for day, pid, name, cin, cout, cin_snap, cout_snap in s.execute(stmt).all():
            last = cout or cin
            dur = (last - cin).total_seconds() if (cin and last) else None
            rows.append({
                "snapshot": cin_snap or cout_snap or "",
                "day": day,
                "person_name": name or (f"Person {str(pid)[:8]}" if pid else "Unknown"),
                "first_in": iso(cin), "last_out": iso(cout) or iso(cin),
                "duration": _fmt_duration(dur),
            })
    return _respond(columns, rows, format, "Attendance")


# ── 2. Group: Headcount, Attendance Compliance ─────────────────────────────
@router.get("/reports/group")
def report_group(day_from: str = Query(...), day_to: str = Query(...),
                 format: str = Query("json"),
                 _: None = Depends(require_service_token)) -> Response:
    columns = ["group", "headcount", "present", "compliance_pct"]
    with session() as s:
        # Total enrolled per group (headcount).
        head = dict(s.execute(
            select(FRSPerson.group_id, func.count())
            .where(FRSPerson.group_id.isnot(None))
            .group_by(FRSPerson.group_id)).all())
        # Distinct persons in this group seen at least once in range (present).
        present_rows = s.execute(
            select(FRSPerson.group_id, func.count(func.distinct(FRSAttendance.person_id)))
            .join(FRSPerson, FRSPerson.id == FRSAttendance.person_id)
            .where(and_(FRSAttendance.day_key >= day_from, FRSAttendance.day_key <= day_to,
                        FRSPerson.group_id.isnot(None)))
            .group_by(FRSPerson.group_id)).all()
        present = dict(present_rows)
        groups = s.execute(select(FRSGroup.id, FRSGroup.name)).all()
        rows = []
        for gid, gname in groups:
            hc = int(head.get(gid, 0))
            pr = int(present.get(gid, 0))
            comp = round(100.0 * pr / hc, 1) if hc else 0.0
            rows.append({"group": gname, "headcount": hc, "present": pr,
                         "compliance_pct": comp})
        rows.sort(key=lambda r: r["headcount"], reverse=True)
    return _respond(columns, rows, format, "Group")


# ── 3. Entry/Exit Mismatch: unpaired vs resolved ───────────────────────────
@router.get("/reports/mismatch")
def report_mismatch(day_from: str = Query(...), day_to: str = Query(...),
                    format: str = Query("json"),
                    _: None = Depends(require_service_token)) -> Response:
    columns = ["snapshot", "person_name", "entry_time", "exit_time", "status"]
    start = naive(datetime.fromisoformat(day_from)) if "T" in day_from else naive(datetime.fromisoformat(day_from + "T00:00:00"))
    end = naive(datetime.fromisoformat(day_to)) if "T" in day_to else naive(datetime.fromisoformat(day_to + "T23:59:59"))
    with session() as s:
        stmt = (select(TransitSession, FRSPerson.full_name)
                .outerjoin(FRSPerson, FRSPerson.id == TransitSession.person_id)
                .where(and_(TransitSession.started_at >= start, TransitSession.started_at <= end))
                .order_by(TransitSession.started_at.desc()))
        rows = []
        for sess, name in s.execute(stmt).all():
            attrs = sess.attributes or {}
            # closed = resolved (paired entry+exit); open/overdue = unpaired/unresolved.
            if sess.status == "closed":
                status = "Resolved"
            elif sess.status == "overdue":
                status = "Unresolved (overdue)"
            else:
                status = "Unpaired (no exit)"
            rows.append({
                "snapshot": attrs.get("entry_snapshot") or attrs.get("face_snapshot")
                or attrs.get("snapshot") or "",
                "person_name": name or attrs.get("person_name")
                or (f"Person {str(sess.person_id)[:8]}" if sess.person_id else "Unknown"),
                "entry_time": iso(sess.started_at),
                "exit_time": iso(sess.ended_at) or "—",
                "status": status,
            })
    return _respond(columns, rows, format, "Entry-Exit Mismatch")


# ── 4. Unknown Attempts: count + snapshots ─────────────────────────────────
@router.get("/reports/unknown")
def report_unknown(day_from: str = Query(...), day_to: str = Query(...),
                   format: str = Query("json"),
                   _: None = Depends(require_service_token),
                   allowed: Optional[list[str]] = Depends(allowed_camera_ids)) -> Response:
    # "confidence" here is the DETECTOR confidence (a face was found) — the match score
    # is always 0 on an Unknown, so showing that read as a confusing "0%".
    columns = ["snapshot", "time", "camera_id", "detected_pct"]
    start = naive(datetime.fromisoformat(day_from + "T00:00:00")) if "T" not in day_from else naive(datetime.fromisoformat(day_from))
    end = naive(datetime.fromisoformat(day_to + "T23:59:59")) if "T" not in day_to else naive(datetime.fromisoformat(day_to))
    with session() as s:
        conds = [FRSEvent.event_type == "face_unknown",
                 FRSEvent.triggered_at >= start, FRSEvent.triggered_at <= end]
        if allowed is not None:
            if not allowed:
                return _respond(columns, [], format, "Unknown Attempts")
            conds.append(FRSEvent.camera_id.in_(allowed))
        stmt = (select(FRSEvent).where(and_(*conds))
                .order_by(FRSEvent.triggered_at.desc()).limit(2000))
        evs = s.execute(stmt).scalars().all()
        rows = []
        for e in evs:
            attrs = e.attributes or {}
            # Prefer the stored detector confidence; fall back to the event confidence
            # only if it's non-zero (older rows had no det_confidence attr).
            det = attrs.get("det_confidence")
            if det is None:
                det = float(e.confidence or 0.0)
            rows.append({
                "snapshot": attrs.get("face_snapshot") or e.snapshot_path or "",
                "time": iso(e.triggered_at),
                "camera_id": e.camera_id,
                "detected_pct": round(float(det) * 100, 1),
            })
    # The UI also wants a total count up top — include it in JSON; exports list rows.
    if (format or "json").lower() == "json":
        return JSONResponse({"columns": columns, "items": rows, "total": len(rows)})
    return _respond(columns, rows, format, "Unknown Attempts")
