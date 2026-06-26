"""Person CRUD (gallery)."""
from __future__ import annotations

import csv
import io
import os
import shutil
import uuid
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import Response, StreamingResponse
from pydantic import ValidationError
from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError

from config import DATA_PATH
from db import session
from deps import require_service_token, purge_person_biometrics
from db.models import FRSPerson, FRSGroup
from schemas import person_dict, PersonCreate, PersonUpdate
from routers.photos import enroll_photo_bytes

router = APIRouter(tags=["persons"])

_IMPORT_MAX_ROWS = 2000
_IMPORT_COLUMNS = [
    "full_name", "external_id", "group", "category", "priority", "gender", "age",
    "department", "designation", "contact_number", "date_of_joining",
    "id_type", "id_number", "validity_start", "validity_end", "auto_remove",
    "photo_file",
]

_COLUMN_ALIASES = {
    "name": "full_name",
    "full name": "full_name",
    "fullname": "full_name",
    "person": "full_name",
    "employee name": "full_name",
    "external id": "external_id",
    "external_id": "external_id",
    "employee id": "external_id",
    "emp id": "external_id",
    "badge": "external_id",
    "badge no": "external_id",
    "group": "group",
    "group name": "group",
    "group_id": "group",
    "group id": "group",
    "category": "category",
    "priority": "priority",
    "gender": "gender",
    "age": "age",
    "department": "department",
    "designation": "designation",
    "profile": "designation",
    "role": "designation",
    "contact": "contact_number",
    "contact number": "contact_number",
    "phone": "contact_number",
    "mobile": "contact_number",
    "date of joining": "date_of_joining",
    "joining date": "date_of_joining",
    "doj": "date_of_joining",
    "id type": "id_type",
    "id_type": "id_type",
    "id number": "id_number",
    "id_number": "id_number",
    "validity start": "validity_start",
    "valid from": "validity_start",
    "validity_start": "validity_start",
    "validity end": "validity_end",
    "valid till": "validity_end",
    "valid until": "validity_end",
    "validity_end": "validity_end",
    "auto remove": "auto_remove",
    "auto_remove": "auto_remove",
    "photo": "photo_file",
    "photo file": "photo_file",
    "photo filename": "photo_file",
    "image": "photo_file",
    "image file": "photo_file",
}


def _clean(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _header_key(v) -> str | None:
    raw = _clean(v).lower().replace("-", " ").replace("_", " ")
    raw = " ".join(raw.split())
    return _COLUMN_ALIASES.get(raw)


def _date_value(v):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    text = _clean(v)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(f"invalid date: {text}") from exc


def _bool_value(v) -> bool:
    if isinstance(v, bool):
        return v
    text = _clean(v).lower()
    return text in ("1", "true", "yes", "y", "on", "auto")


def _int_value(v, default=0) -> int:
    text = _clean(v)
    if not text:
        return default
    return int(float(text))


def _split_files(v) -> list[str]:
    text = _clean(v)
    if not text:
        return []
    out: list[str] = []
    for part in text.replace("\n", ",").replace(";", ",").split(","):
        name = part.strip().strip("\"'")
        if name:
            out.append(os.path.basename(name))
    return out


def _rows_from_csv(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return _rows_from_table(list(reader))


def _rows_from_xlsx(data: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(500, "Excel import requires openpyxl in the FRS image") from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    return _rows_from_table(list(ws.iter_rows(values_only=True)))


def _rows_from_table(table) -> list[dict]:
    header = None
    start_idx = 0
    for idx, row in enumerate(table):
        mapped = [_header_key(c) for c in row]
        if any(mapped):
            header = mapped
            start_idx = idx + 1
            break
    if not header:
        raise HTTPException(400, "import sheet header not found")
    rows: list[dict] = []
    for idx, row in enumerate(table[start_idx:], start=start_idx + 1):
        if len(rows) >= _IMPORT_MAX_ROWS:
            raise HTTPException(413, f"import limit is {_IMPORT_MAX_ROWS} rows")
        data: dict = {"_row": idx}
        empty = True
        for col_idx, key in enumerate(header):
            if not key:
                continue
            val = row[col_idx] if col_idx < len(row) else None
            if _clean(val):
                empty = False
            data[key] = val
        if not empty:
            rows.append(data)
    return rows


def _parse_import_sheet(file_name: str, data: bytes) -> list[dict]:
    lower = file_name.lower()
    if lower.endswith(".csv"):
        return _rows_from_csv(data)
    if lower.endswith(".xlsx"):
        return _rows_from_xlsx(data)
    raise HTTPException(415, "upload an .xlsx or .csv import sheet")


def _group_id_for(s, value) -> str | None:
    text = _clean(value)
    if not text:
        return None
    by_id = s.get(FRSGroup, text)
    if by_id:
        return by_id.id
    group = s.scalar(select(FRSGroup).where(func.lower(FRSGroup.name) == text.lower()))
    if not group:
        raise ValueError(f"group not found: {text}")
    return group.id


def _payload_from_row(s, row: dict) -> tuple[dict, list[str]]:
    attrs = {}
    gender = _clean(row.get("gender")).lower()
    if gender:
        attrs["gender"] = gender
    age = _clean(row.get("age"))
    if age:
        attrs["age"] = _int_value(age)
    payload = {
        "full_name": _clean(row.get("full_name")),
        "external_id": _clean(row.get("external_id")) or None,
        "group_id": _group_id_for(s, row.get("group")),
        "category": _clean(row.get("category")) or "standard",
        "priority": _int_value(row.get("priority"), 0),
        "attributes": attrs or None,
        "department": _clean(row.get("department")) or None,
        "designation": _clean(row.get("designation")) or None,
        "contact_number": _clean(row.get("contact_number")) or None,
        "date_of_joining": _date_value(row.get("date_of_joining")),
        "id_type": _clean(row.get("id_type")) or None,
        "id_number": _clean(row.get("id_number")) or None,
        "validity_start": _date_value(row.get("validity_start")),
        "validity_end": _date_value(row.get("validity_end")),
        "auto_remove": _bool_value(row.get("auto_remove")),
    }
    PersonCreate(**payload)
    return payload, _split_files(row.get("photo_file"))


def _apply_person_payload(person: FRSPerson, payload: dict) -> None:
    for key, value in payload.items():
        setattr(person, key, value)


@router.get("/persons")
def list_persons(limit: int = Query(50, ge=1, le=200), offset: int = Query(0, ge=0),
                 search: Optional[str] = None, group_id: Optional[str] = None,
                 category: Optional[str] = None, _: None = Depends(require_service_token)) -> dict:
    with session() as s:
        conds = []
        if search:
            like = f"%{search.strip()}%"
            conds.append(or_(FRSPerson.full_name.ilike(like), FRSPerson.external_id.ilike(like)))
        if group_id:
            conds.append(FRSPerson.group_id == group_id)
        if category:
            conds.append(FRSPerson.category == category)
        cq = select(func.count(FRSPerson.id))
        rq = select(FRSPerson)
        for c in conds:
            cq = cq.where(c); rq = rq.where(c)
        total = int(s.scalar(cq) or 0)
        rows = s.execute(rq.order_by(FRSPerson.created_at.desc()).limit(limit).offset(offset)).scalars().all()
        return {"items": [person_dict(p) for p in rows], "total": total, "limit": limit, "offset": offset}


@router.post("/persons", status_code=201)
def create_person(body: PersonCreate, _: None = Depends(require_service_token)) -> dict:
    with session() as s:
        if body.group_id and not s.get(FRSGroup, body.group_id):
            raise HTTPException(400, "group not found")
        p = FRSPerson(
            full_name=body.full_name, external_id=body.external_id,
            group_id=body.group_id, category=body.category,
            priority=body.priority, attributes=body.attributes,
            department=body.department, designation=body.designation,
            contact_number=body.contact_number, date_of_joining=body.date_of_joining,
            id_type=body.id_type, id_number=body.id_number,
            validity_start=body.validity_start, validity_end=body.validity_end,
            auto_remove=bool(body.auto_remove),
        )
        s.add(p); s.commit(); s.refresh(p)
        return person_dict(p)


@router.get("/persons/import-template")
def persons_import_template(_: None = Depends(require_service_token)):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(500, "Excel template requires openpyxl in the FRS image") from exc
    wb = Workbook()
    ws = wb.active
    ws.title = "Persons"
    ws.append(_IMPORT_COLUMNS)
    ws.append([
        "Asha Sharma", "EMP001", "GVD", "standard", 0, "female", 31,
        "Operations", "Manager", "+919999999999", "2026-01-15",
        "Company ID", "CID-001", "2026-01-01", "2026-06-30", "no",
        "asha.jpg",
    ])
    ws.freeze_panes = "A2"
    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 28)
        ws.column_dimensions[col[0].column_letter].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="frs_persons_import_template.xlsx"'},
    )


@router.post("/persons/import")
async def import_persons(
    sheet: UploadFile = File(...),
    photos: Optional[list[UploadFile]] = File(default=None),
    update_existing: bool = Form(default=True),
    _: None = Depends(require_service_token),
) -> dict:
    data = await sheet.read()
    if not data:
        raise HTTPException(400, "empty import sheet")
    rows = _parse_import_sheet(sheet.filename or "", data)
    uploads: dict[str, tuple[bytes, str | None]] = {}
    for f in photos or []:
        if not f.filename:
            continue
        uploads[os.path.basename(f.filename).lower()] = (await f.read(), f.content_type)

    results: list[dict] = []
    created = updated = skipped = photos_enrolled = photos_failed = 0
    for row in rows:
        row_no = int(row.get("_row") or 0)
        full_name = _clean(row.get("full_name"))
        if not full_name:
            skipped += 1
            results.append({"row": row_no, "status": "skipped", "error": "full_name is required"})
            continue
        try:
            with session() as s:
                payload, photo_names = _payload_from_row(s, row)
                person = None
                if payload.get("external_id"):
                    person = s.scalar(select(FRSPerson).where(FRSPerson.external_id == payload["external_id"]))
                action = "created"
                if person is not None:
                    if not update_existing:
                        skipped += 1
                        results.append({
                            "row": row_no, "status": "skipped",
                            "person_id": person.id, "external_id": person.external_id,
                            "error": "external_id already exists",
                        })
                        continue
                    _apply_person_payload(person, payload)
                    action = "updated"
                else:
                    person = FRSPerson(**payload)
                    s.add(person)
                try:
                    s.commit()
                except IntegrityError as exc:
                    s.rollback()
                    raise ValueError("external_id already exists") from exc
                s.refresh(person)
                person_id = person.id

            photo_results = []
            for name in photo_names:
                blob = uploads.get(name.lower())
                if blob is None:
                    photos_failed += 1
                    photo_results.append({"file": name, "status": "missing"})
                    continue
                try:
                    ph = enroll_photo_bytes(person_id, blob[0], content_type=blob[1])
                    if ph.get("status") == "enrolled":
                        photos_enrolled += 1
                    else:
                        photos_failed += 1
                    photo_results.append({"file": name, "status": ph.get("status"), "photo_id": ph.get("id"), "error": ph.get("error")})
                except Exception as exc:  # noqa: BLE001
                    photos_failed += 1
                    photo_results.append({"file": name, "status": "failed", "error": str(exc)})

            if action == "created":
                created += 1
            else:
                updated += 1
            results.append({
                "row": row_no, "status": action, "person_id": person_id,
                "external_id": payload.get("external_id"), "full_name": payload.get("full_name"),
                "photos": photo_results,
            })
        except (ValidationError, ValueError, HTTPException) as exc:
            skipped += 1
            detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
            results.append({"row": row_no, "status": "failed", "full_name": full_name, "error": detail})
        except Exception as exc:  # noqa: BLE001
            skipped += 1
            results.append({"row": row_no, "status": "failed", "full_name": full_name, "error": str(exc)})

    return {
        "total": len(rows),
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "photos_enrolled": photos_enrolled,
        "photos_failed": photos_failed,
        "results": results,
    }


@router.get("/persons/{person_id}")
def get_person(person_id: str, _: None = Depends(require_service_token)) -> dict:
    with session() as s:
        p = s.get(FRSPerson, person_id)
        if not p:
            raise HTTPException(404, "person not found")
        return person_dict(p)


@router.put("/persons/{person_id}")
def update_person(person_id: str, body: PersonUpdate, _: None = Depends(require_service_token)) -> dict:
    with session() as s:
        p = s.get(FRSPerson, person_id)
        if not p:
            raise HTTPException(404, "person not found")
        if body.group_id and not s.get(FRSGroup, body.group_id):
            raise HTTPException(400, "group not found")
        # Only patch fields the client actually sent (exclude_unset).
        for k, v in body.model_dump(exclude_unset=True).items():
            setattr(p, k, v)
        s.commit(); s.refresh(p)
        return person_dict(p)


@router.delete("/persons/{person_id}", status_code=204)
def delete_person(person_id: str, _: None = Depends(require_service_token)):
    """Right-to-erasure: purge ALL biometric traces of this person — gallery
    vectors, live-sighting vectors, snapshot files, events, attendance, photos,
    and the on-disk photo directory — in one transaction (GDPR/BIPA)."""
    id_key = None
    with session() as s:
        p = s.get(FRSPerson, person_id)
        if not p:
            raise HTTPException(404, "person not found")
        id_key = p.id_file_key
        purge_person_biometrics(s, person_id)   # events + attendance + photos + vectors + snapshot files
        s.delete(p)
        s.commit()
    photo_dir = DATA_PATH / "persons" / person_id
    if photo_dir.exists():
        shutil.rmtree(photo_dir, ignore_errors=True)
    if id_key:
        _delete_id_object(id_key)
    return Response(status_code=204)


# ── ID document (image/PDF in the object store) ──────────────────────────────
_ID_MAX_BYTES = 15 * 1024 * 1024
_ID_TYPES = {
    "image/jpeg": "jpg", "image/png": "png", "image/webp": "webp",
    "application/pdf": "pdf",
}


def _delete_id_object(key: str) -> None:
    try:
        from vizor_sdk.objectstore import default_store
        default_store().delete(key)
    except Exception:  # noqa: BLE001 — best-effort
        pass


@router.post("/persons/{person_id}/id-document")
async def upload_id_document(
    person_id: str,
    file: UploadFile = File(...),
    _: None = Depends(require_service_token),
) -> dict:
    """Store a person's government/company ID (image or PDF) in the object store and
    record its key. Replaces any previous document."""
    ct = (file.content_type or "").split(";")[0].strip()
    ext = _ID_TYPES.get(ct)
    if not ext:
        raise HTTPException(415, "ID document must be JPG, PNG, WEBP or PDF")
    data = await file.read()
    if not data:
        raise HTTPException(400, "empty upload")
    if len(data) > _ID_MAX_BYTES:
        raise HTTPException(413, "ID document exceeds the 15 MB limit")
    from vizor_sdk.objectstore import default_store
    store = default_store()
    key = f"frs/ids/{person_id}/{uuid.uuid4().hex}.{ext}"
    with session() as s:
        p = s.get(FRSPerson, person_id)
        if not p:
            raise HTTPException(404, "person not found")
        old = p.id_file_key
        store.put(key, data, ct)
        p.id_file_key = key
        s.commit()
        s.refresh(p)
        result = person_dict(p)
    if old and old != key:
        _delete_id_object(old)
    return result


@router.get("/persons/{person_id}/id-document")
def get_id_document(person_id: str, _: None = Depends(require_service_token)):
    """Serve the stored ID document bytes (proxied — the object store isn't
    browser-reachable)."""
    with session() as s:
        p = s.get(FRSPerson, person_id)
        if not p:
            raise HTTPException(404, "person not found")
        key = p.id_file_key
    if not key:
        raise HTTPException(404, "no ID document")
    from vizor_sdk.objectstore import default_store
    try:
        data = default_store().get(key)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(404, "ID document not found") from exc
    media = "application/pdf" if key.endswith(".pdf") else "image/jpeg"
    if key.endswith(".png"):
        media = "image/png"
    elif key.endswith(".webp"):
        media = "image/webp"
    return Response(content=data, media_type=media)


@router.delete("/persons/{person_id}/id-document", status_code=204)
def delete_id_document(person_id: str, _: None = Depends(require_service_token)):
    with session() as s:
        p = s.get(FRSPerson, person_id)
        if not p:
            raise HTTPException(404, "person not found")
        key = p.id_file_key
        p.id_file_key = None
        s.commit()
    if key:
        _delete_id_object(key)
    return Response(status_code=204)
